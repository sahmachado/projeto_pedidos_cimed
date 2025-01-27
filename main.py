import pandas as pd
from openpyxl import load_workbook
import tkinter as tk
from tkinter import ttk
import datetime as dt
from tkinter import messagebox
from PIL import Image, ImageTk

class style:
    def style():
        style = ttk.Style()

        style.theme_use('clam')
        style.configure("Menu.TButton", padding=(10, 5), width=21, font=("Arial", 10), background='#ffcd00')
        style.map("Menu.TButton", 
                background=[("active", "#3883ba"), ("pressed", "#3883ba")],
                foreground=[('active', 'white')])

        style.configure('p.TButton',background='#333',foreground = 'white')
        style.map('p.TButton',
                background=[("active", "#0056b3"), ("pressed", "#0056b3")],
                foreground=[('active', 'white')])
        style.configure('ed.TButton',background='#333',foreground = 'white')
        style.map('ed.TButton',
                background=[("active", "#87CEEB"), ("pressed", "#87CEEB")],
                foreground=[('active', 'black')])
        style.configure('s.TButton',background='#333',foreground = 'white')
        style.map('s.TButton',
                background=[("active", "#1e7e34"), ("pressed", "#1e7e34")],
                foreground=[('active', 'white')])
        style.configure('c.TButton',background='#333',foreground = 'white')
        style.map('c.TButton',
                background=[("active", "#A9A9A9"), ("pressed", '#A9A9A9')],
                foreground=[('active', 'black')])
        style.configure('ex.TButton',background='#333',foreground = 'white')
        style.map('ex.TButton',
                background=[("active", "#bd2130"), ("pressed", "#bd2130")],
                foreground=[('active', 'white')])

        style.configure("ActiveMenu.TButton", padding=(10, 5), width=21, font=("Arial", 10), background='#00cccc')
        style.map("ActiveMenu.TButton", 
                background=[("active", "#3883ba"), ("pressed", "#3883ba")],
                foreground=[('active', 'white')])

        style.configure("MenuFrame.TFrame", background='#f0f0f0')
        style.configure("TLabel", background="#f0f0f0")
        style.configure("TEntry", padding=5)
        style.configure('C.TFrame', background='#f0f0f0f0')

class limpar_campos:
    def limpar_pesquisa():
        num_pedido.delete(0, tk.END)
        comprador.config(state='normal')
        comprador.delete(0, tk.END)
        comprador.config(state='readonly')
        fornecedor.config(state='normal')
        fornecedor.delete(0, tk.END)
        fornecedor.config(state='readonly')
        material.config(state='normal')
        material.delete(0, tk.END)
        material.config(state='readonly')
        item_combo.delete(0, tk.END)
        remessa.config(state='normal')
        remessa.delete(0, tk.END)
        remessa.config(state='readonly')
        status.config(state='normal')
        status.delete(0, tk.END)
        status.config(state='readonly')
        follow_up.config(state='normal')
        follow_up.delete(0, tk.END)
        follow_up.config(state='readonly')
        
class processarpedido:
    def atualizar_pedido():
        """Procura um valor na primeira coluna e retorna a linha."""
        pedido = num_pedido.get()
        item = item_combo.get()
        codigo = processarpedido.gerar_codigo(pedido,item)
        if pedido != '':
            if follow_up.cget("state") == "normal":
                if follow_up.get() =='':
                    r =messagebox.askyesno('Follow-Up','O campo FOLLOW-UP está vazio, deseja continuar?')
                    follow_up.focus_set()
                    if r == True:
                        arquivo = load_workbook('base teste.xlsx')
                        aba_atual = arquivo.active
                        for linha, linha_dados in enumerate(aba_atual.iter_rows(min_row=1), start=1):
                            if linha_dados[0].value == int(codigo): # Verifica apenas a primeira célula da linha
                                aba_atual.cell(linha,column=5).value = comprador.get()
                                aba_atual.cell(linha,column=6).value = remessa.get()
                                aba_atual.cell(linha,column=8).value = follow_up.get()
                                arquivo.save('base teste.xlsx')
                                limpar_campos.limpar_pesquisa()
                                processarpedido.save_pedido
            else:
                messagebox.showwarning("Erro", "Habilite o modo edição")
        else: 
            messagebox.showwarning("Erro", "Insira um número de pedido para edição")

    def validar_pedido(pedido):
        pedido = str(pedido)
        if len(pedido) != 10:
            status = 'pedido invalido'
        elif pedido[:2] != '43' and pedido[:2] != '45' and pedido[:2] != '46':
            status = 'pedido invalido'
        else:
            status = 'pedido valido'
        return status

    def gerar_codigo(pedido,item):
        if pedido[:2] == '45':
            final_pedido = pedido[4:]
        elif pedido[:2] == '46':
            final_pedido = pedido[5:]
        elif pedido[:2] == '43':
            final_pedido = pedido[6:]
        codigo = f'{final_pedido}{item}'

        return codigo

    def ao_selecionar_item(event):
        
        base = pd.read_excel('base teste.xlsx', dtype={"Data de Remessa": "datetime64[ns]"})

        codigo = processarpedido.gerar_codigo(num_pedido.get(),item_combo.get())

        material_filto = base.loc[base['Codigo'] == int(float(codigo)), 'Material']
        material.config(state='normal')
        material.delete(0, tk.END)
        material.insert(0,material_filto.values[0])
        material.config(state='readonly')

        remessa_valor = base.loc[base['Codigo'] == int(float(codigo)), 'Data de Remessa']
        remessa_valor = dt.datetime.strptime(remessa_valor.iloc[0].strftime("%Y-%m-%d"), "%Y-%m-%d").date()
        # remessa_valor = pd.to_datetime(remessa_valor, format="%Y-%m-%d")
        remessa.config(state='normal')
        remessa.delete(0, tk.END)
        remessa.insert(0, remessa_valor.strftime("%d/%m/%Y"))
        remessa.config(state='readonly')

        if remessa_valor >= dt.date.today():
            status.config(state='normal')
            status.delete(0, tk.END)
            status.insert(0, 'No Prazo')
            status.config(state='readonly')
        else:
            status.config(state='normal')
            status.delete(0, tk.END)
            status.insert(0, 'Em atraso')
            status.config(state='readonly')

        follow_up_valor = base.loc[base['Codigo'] == int(float(codigo)), 'Follow-up']
        
        if pd.isnull(follow_up_valor.iloc[0]):  # ou pd.isna(...)
            follow_up.config(state='normal')
            follow_up.delete(0, tk.END)
            follow_up.insert(0, 'Sem previsão')
            follow_up.config(state='readonly')
        else:
            follow_up.config(state='normal')
            follow_up.delete(0, tk.END)
            follow_up.insert(0, follow_up_valor.iloc[0])
            follow_up.config(state='readonly')

    def pesquisar_pedido():
        # try:
            # Obter o número do pedido inserido
            pedido = int(num_pedido.get())
            if pedido == '':
                messagebox.showerror('Erro','Insira um pedido')
                num_pedido.focus_set()
            else:
                status_pedido = processarpedido.validar_pedido(pedido)
                
                if status_pedido == 'pedido invalido':
                    messagebox.showerror("Erro", "Por favor, insira um número de pedido válido.")
                else:
                    # Carregar a base de dados
                    base = pd.read_excel('base teste.xlsx',sheet_name='base',dtype = {"Data de Remessa": "datetime64[ns]", "Data do Pedido": "datetime64[ns]"})

                    # Filtrar os itens do pedido
                    itens_pedido = base.loc[base['Pedido'] ==  pedido, 'Item'].tolist()
                    itens = []
                    for item in itens_pedido:
                        itens.append(int(float(item)))
                    # Atualizar os valores da ComboBox
                    if itens:
                        item_combo['values'] = itens
                        item_combo.current(0)  # Selecionar o primeiro item automaticamente
                    else:
                        item_combo['values'] = ["Nenhum item encontrado"]
                        item_combo.set("Nenhum item encontrado")

                    # Filtrar comprador com base no pedido
                    comprador_valor = base.loc[base['Pedido'] ==  pedido, 'Comprador']
                    if not comprador_valor.empty:
                        comprador.config(state='normal')
                        comprador.delete(0, tk.END)
                        comprador.insert(0, comprador_valor.values[0])
                        comprador.config(state='readonly')
                    else:
                        messagebox.showwarning("Aviso", "Comprador não encontrado.")
                        return

                    criacao_valor = base.loc[base['Pedido'] ==  pedido, 'Data do Pedido']
                    criacao_valor = dt.datetime.strptime(criacao_valor.iloc[0].strftime("%Y-%m-%d"), "%Y-%m-%d").date()
                    criacao_valor = pd.to_datetime(criacao_valor, format="%Y-%m-%d")
                    criacao.config(state='normal')
                    criacao.delete(0, tk.END)
                    criacao.insert(0, criacao_valor.strftime("%d/%m/%Y"))
                    criacao.config(state='readonly')

                    fornecedor_valor = base.loc[base['Pedido'] ==  pedido, 'Fornecedor']
                    if not fornecedor_valor.empty:
                        fornecedor.config(state='normal')
                        fornecedor.delete(0, tk.END)
                        fornecedor.insert(0, fornecedor_valor.values[0])
                        fornecedor.config(state='readonly')
                    else:
                        messagebox.showwarning("Aviso", "Fornecedor não encontrado.")
                        return
                    
                    # Encontrar primeiro item do pedido
                    codigo = processarpedido.gerar_codigo(str( pedido),itens[0])

                    material_filto = base.loc[base['Codigo'] == int(float(codigo)), 'Material']
                    material.config(state='normal')
                    material.delete(0, tk.END)
                    material.insert(0,material_filto.values[0])
                    material.config(state='readonly')

                    remessa_valor = base.loc[base['Codigo'] == int(float(codigo)), 'Data de Remessa']
                    remessa_valor = dt.datetime.strptime(remessa_valor.iloc[0].strftime("%Y-%m-%d"), "%Y-%m-%d").date()
                    # remessa_valor = dt.datetime.strptime(remessa_valor, "%d/%m/%Y").date()
                    remessa_valor = pd.to_datetime(remessa_valor, format="%Y-%m-%d")
                    remessa.config(state='normal')
                    remessa.delete(0, tk.END)
                    remessa.insert(0, remessa_valor.strftime("%d/%m/%Y"))
                    remessa.config(state='readonly')

                    if remessa_valor.date() >= dt.date.today():
                        status.config(state='normal')
                        status.delete(0, tk.END)
                        status.insert(0, 'No Prazo')
                        status.config(state='readonly')
                    else:
                        status.config(state='normal')
                        status.delete(0, tk.END)
                        status.insert(0, 'Em atraso')
                        status.config(state='readonly')

                    follow_up_valor = base.loc[base['Codigo'] == int(float(codigo)), 'Follow-up']
                    if pd.isnull(follow_up_valor.iloc[0]):  # ou pd.isna(...)
                        follow_up.config(state='normal')
                        follow_up.delete(0, tk.END)
                        follow_up.insert(0, 'Sem previsão')
                        follow_up.config(state='readonly')
                    else:
                        follow_up.config(state='normal')
                        follow_up.delete(0, tk.END)
                        follow_up.insert(0, follow_up_valor.iloc[0])
                        follow_up.config(state='readonly')

        # except ValueError:
        #     messagebox.showerror("Erro", "Por favor, insira um número de pedido válido.")
        # except FileNotFoundError:
        #     messagebox.showerror("Erro", "Arquivo 'base teste.xlsx' não encontrado.")
        # except Exception as e:
        #     messagebox.showerror("Erro", f"Ocorreu um erro inesperado: {e}")

    def editar_pedido():
        pedido = material.get()
        if pedido != '':
            comprador.config(state='normal')
            remessa.config(state='normal')
            follow_up.config(state='normal')
            modo_edicao['text'] = 'Modo Edição'
            messagebox.showinfo('Edição','Modo edição ativado')
        else: 
            messagebox.showwarning("Erro", "Insira um número de pedido para edição")

    def save_pedido():
        messagebox.showinfo("Salvar", "Pedido salvo com sucesso!")

    def cancel_pedido():
        material.config(state='readonly')
        fornecedor.config(state='readonly')
        remessa.config(state='readonly')
        status.config(state='readonly')
        follow_up.config(state='readonly')
        modo_edicao['text'] = ''
        messagebox.showwarning("Cancelar", "Operação cancelada!")

    def delete_pedido():

        arquivo = load_workbook('base teste.xlsx')
        aba_atual = arquivo.active
        pedido = num_pedido.get()
        item = item_combo.get()
        codigo = processarpedido.gerar_codigo(pedido,item)
        if pedido != '':
            r =messagebox.askyesno('Excluir',f'Deseja excluir o item {item} do pedido {pedido}?')
            if r == True:
                arquivo = load_workbook('base teste.xlsx')
                aba_atual = arquivo.active
                for linha, linha_dados in enumerate(aba_atual.iter_rows(min_row=1), start=1):
                    if linha_dados[0].value == int(codigo): 
                        aba_atual.delete_rows(linha)
                        arquivo.save('base teste.xlsx')
                        messagebox.showerror("Excluir", "Pedido excluído!")
        else:
            messagebox.showwarning('Erro','Selecione um pedido')

class processarFornecedor:
    
    def pesquisa_fornecedor():
        def preencher_campos(fornecedor,codigo,nome,email):
            razao_fornecedor.config(state='normal')
            razao_fornecedor.delete(0, tk.END)
            razao_fornecedor.insert(0, fornecedor)
            razao_fornecedor.config(state='readonly')

            codigo_fornecedor.delete(0, tk.END)
            codigo_fornecedor.insert(0, int(float(codigo)))
  
            nome_fornecedor.delete(0, tk.END)
            nome_fornecedor.insert(0, nome)
        
            email_fornecedor.config(state='normal')
            email_fornecedor.delete(0, tk.END)
            email_fornecedor.insert(0, email)
            email_fornecedor.config(state='readonly')

        base = pd.read_excel('base teste.xlsx',sheet_name='fornecedores')
        fornecedor = nome_fornecedor.get()
        fornecedor = fornecedor.upper()
        nome_resultado = base.loc[base['Nome'] ==  fornecedor, 'Fornecedor']
        if nome_fornecedor.get() == '' and codigo_fornecedor.get() =='':
            messagebox.showwarning('Erro', 'Preencha o campo de NOME ou CODIGO para pesquisa')
            nome_fornecedor.focus_set()
        else:
            if fornecedor == '':
                codigo = int(codigo_fornecedor.get())
                codigo_resultado = base.loc[base['Codigo'] ==  codigo, 'Fornecedor']
                
            if not nome_resultado.empty:
                fornecedor_result = base.loc[base['Nome'] ==  fornecedor, 'Fornecedor']
                codigo_result = base.loc[base['Nome'] ==  fornecedor, 'Codigo']
                email_result = base.loc[base['Nome'] ==  fornecedor, 'Email']
                preencher_campos(fornecedor_result.values[0],codigo_result.values[0],fornecedor,email_result.values[0])
            elif not codigo_resultado.empty:
                fornecedor_result = base.loc[base['Codigo'] ==  codigo, 'Fornecedor']
                nome_result = base.loc[base['Codigo'] ==  codigo, 'Nome']
                email_result = base.loc[base['Codigo'] ==  codigo, 'Email']
                preencher_campos(fornecedor_result.values[0],codigo,nome_result.values[0],email_result.values[0])
            else:
                messagebox.showerror('Erro','Fornecedor não encontrado')
    
    def limpar_campos():
        razao_fornecedor.config(state='normal')
        razao_fornecedor.delete(0, tk.END)
        razao_fornecedor.config(state='readonly')
        codigo_fornecedor.delete(0, tk.END)
        nome_fornecedor.delete(0, tk.END)
        email_fornecedor.config(state='normal')
        email_fornecedor.delete(0, tk.END)
        email_fornecedor.config(state='readonly')


class paginas:
    def pagina_pedidos():
        global num_pedido, material, item_combo, fornecedor,remessa,status,follow_up,modo_edicao,comprador,criacao
        
        titulo = ttk.Label(content_frame, text='Consulta e Atualização de pedidos', font=("Arial", 15),background='#DCDAD5')
        titulo.grid(row=0, column=1,columnspan=2, padx=(15,5), pady=(15, 5))

        modo_edicao = ttk.Label(content_frame, text='', font=("Arial", 9),background='#DCDAD5',foreground='red')
        modo_edicao.grid(row=0, column=0,sticky="w", padx=(15,5), pady=(15, 5))

        ttk.Label(content_frame, text='Nº do Pedido', font=("Arial", 11),background='#DCDAD5').grid(row=1, column=0, sticky="w", padx=(15,5), pady=(15, 5))
        num_pedido = ttk.Entry(content_frame, width=25)
        num_pedido.grid(row=1, column=1, padx=5, pady=(15, 5),sticky='w')
        num_pedido.focus_set() #Coloca

        ttk.Label(content_frame, text='Fornecedor', font=("Arial", 11),background='#DCDAD5').grid(row=1, column=2, sticky="w", padx=(15,5), pady=5)
        fornecedor = ttk.Entry(content_frame, width=40,state="readonly")
        fornecedor.grid(row=1, column=3, padx=5, pady=5,sticky='w')

        ttk.Label(content_frame, text='Comprador', font=("Arial", 11),background='#DCDAD5').grid(row=2, column=0, sticky="w", padx=(15,5), pady=(15, 5))
        comprador = ttk.Entry(content_frame, width=20)
        comprador.grid(row=2, column=1, padx=5, pady=(15, 5),sticky='w')

        ttk.Label(content_frame, text='Data de Criação', font=("Arial", 11),background='#DCDAD5').grid(row=2, column=2, sticky="w", padx=(15,5), pady=(15, 5))
        criacao = ttk.Entry(content_frame, width=20)
        criacao.grid(row=2, column=3, padx=5, pady=(15, 5),sticky='w')

        ttk.Label(content_frame, text='Item do Pedido', font=("Arial", 11),background='#DCDAD5').grid(row=3, column=0, sticky="w", padx=(15,5), pady=5)
        item_combo = ttk.Combobox(content_frame, width=19)
        item_combo['values'] = ["Item Indisponivel"]
        item_combo.set("Item Indisponivel")
        item_combo.grid(row=3, column=1, padx=5, pady=5,sticky='nsw')
        item_combo.current(0)
        item_combo.bind("<<ComboboxSelected>>", processarpedido.ao_selecionar_item)

        ttk.Label(content_frame, text='Material', font=("Arial", 11),background='#DCDAD5').grid(row=3, column=2, sticky="w", padx=(15,5), pady=5)
        material = ttk.Entry(content_frame, width=40,state="readonly")
        material.grid(row=3, column=3, padx=5, pady=5,sticky='w')

        ttk.Label(content_frame, text='Data de Remessa', font=("Arial", 11),background='#DCDAD5').grid(row=4, column=0, sticky="w", padx=(15,5), pady=5)
        remessa = ttk.Entry(content_frame, width=20)
        remessa.grid(row=4, column=1, padx=5, pady=5,sticky='w')

        ttk.Label(content_frame, text='Status', font=("Arial", 11),background='#DCDAD5').grid(row=4, column=2, sticky="w", padx=(15,5), pady=5)
        status = ttk.Entry(content_frame, width=20,state="readonly")
        status.grid(row=4, column=3, padx=5, pady=5,sticky='w')

        ttk.Label(content_frame, text='Follow Up', font=("Arial", 11),background='#DCDAD5').grid(row=5, column=0, sticky="w", padx=(15,5), pady=5)
        follow_up = ttk.Entry(content_frame, width=40)
        follow_up.grid(row=5, column=1,columnspan=1, padx=5, pady=5,sticky='nsew')

        # Botões com espaçamento controlado
        botoes_frame = ttk.Frame(content_frame)
        botoes_frame.grid(row=6, column=0, columnspan=5, pady=20)

        button.button_pedido(botoes_frame)
        
    def pagina_atrasados():
        for i in range(5): 
            content_frame.columnconfigure(i, weight=1)

            ttk.Label(content_frame, text='Pedido', font=("Arial", 11), background='#DCDAD5').grid(row=0, column=0, sticky="ew", padx=(15, 5), pady=(15, 5))
            ttk.Label(content_frame, text='Item', font=("Arial", 11), background='#DCDAD5').grid(row=0, column=1, sticky="ew", padx=(15, 5), pady=(15, 5))
            ttk.Label(content_frame, text='Fornecedor', font=("Arial", 11), background='#DCDAD5').grid(row=0, column=2, sticky="ew", padx=(15, 5), pady=(15, 5))
            ttk.Label(content_frame, text='Comprador', font=("Arial", 11), background='#DCDAD5').grid(row=0, column=3, sticky="ew", padx=(15, 5), pady=(15, 5))
            ttk.Label(content_frame, text='Data de Remessa', font=("Arial", 11), background='#DCDAD5').grid(row=0, column=4, sticky="ew", padx=(15, 5), pady=(15, 5))

    def pagina_cadastro():
        global razao_fornecedor,codigo_fornecedor,nome_fornecedor,email_fornecedor

        ttk.Label(content_frame, text='Razão Social', font=("Arial", 11),background='#DCDAD5').grid(row=0, column=0, sticky="w", padx=(15,5), pady=(15, 5))
        razao_fornecedor = ttk.Entry(content_frame, width=50,state='readonly')
        razao_fornecedor.grid(row=0, column=1, padx=5, pady=(15, 5),sticky='w',)

        ttk.Label(content_frame, text='Código SAP', font=("Arial", 11),background='#DCDAD5').grid(row=0, column=2, sticky="w", padx=(15,5), pady=(15, 5))
        codigo_fornecedor = ttk.Entry(content_frame, width=20)
        codigo_fornecedor .grid(row=0, column=3, padx=5, pady=(15, 5),sticky='w',)

        ttk.Label(content_frame, text='Nome do Fornecedor', font=("Arial", 11),background='#DCDAD5').grid(row=1, column=0, sticky="w", padx=(15,5), pady=(15, 5))
        nome_fornecedor = ttk.Entry(content_frame, width=30)
        nome_fornecedor.grid(row=1, column=1, padx=5, pady=(15, 5),sticky='w',)

        ttk.Label(content_frame, text='E-mail Fornecedor', font=("Arial", 11),background='#DCDAD5').grid(row=2, column=0, sticky="w", padx=(15,5), pady=(15, 5))
        email_fornecedor = ttk.Entry(content_frame, width=50,state='readonly')
        email_fornecedor.grid(row=2, column=1, padx=5, pady=(15, 5),sticky='w',)
    
        botoes_frame = ttk.Frame(content_frame)
        botoes_frame.grid(row=6, column=0, columnspan=4, pady=20)

        button.button_fornecedor(botoes_frame)

    def pagina_follow_up():
        pass

class button:
    def button_pedido(botoes_frame):
        ttk.Button(botoes_frame, text="Pesquisar", command=processarpedido.pesquisar_pedido, width=15,style='p.TButton').grid(row=0, column=0, padx=10)
        ttk.Button(botoes_frame, text="Editar", command=processarpedido.editar_pedido, width=15,style='ed.TButton').grid(row=0, column=1, padx=10)
        ttk.Button(botoes_frame, text="Salvar", command=processarpedido.atualizar_pedido, width=15,style='s.TButton').grid(row=0, column=2, padx=10)
        ttk.Button(botoes_frame, text="Cancelar", command=processarpedido.cancel_pedido, width=15,style='c.TButton').grid(row=0, column=3, padx=10)
        ttk.Button(botoes_frame, text="Excluir", command=processarpedido.delete_pedido, width=15,style='ex.TButton').grid(row=0, column=4, padx=10)
    
    def button_fornecedor(botoes_frame):
        ttk.Button(botoes_frame, text="Pesquisar", command=processarFornecedor.pesquisa_fornecedor, width=15,style='p.TButton').grid(row=0, column=0, padx=10)
        ttk.Button(botoes_frame, text="Editar", command=processarpedido.editar_pedido, width=15,style='ed.TButton').grid(row=0, column=1, padx=10)
        ttk.Button(botoes_frame, text="Salvar", command=processarpedido.save_pedido, width=15,style='s.TButton').grid(row=0, column=2, padx=10)
        ttk.Button(botoes_frame, text="Cancelar", command=processarpedido.cancel_pedido, width=15,style='c.TButton').grid(row=0, column=3, padx=10)
        ttk.Button(botoes_frame, text="Limpar", command=processarFornecedor.limpar_campos, width=15,style='ex.TButton').grid(row=0, column=4, padx=10)

def show_page(page):
    

    for widget in content_frame.winfo_children():
        widget.destroy()

    if page == "pedidos":
        paginas.pagina_pedidos()
    
    elif page == 'atrasados':
        paginas.pagina_atrasados()

    elif page == 'fornecedor':
        pass
    
    elif page == 'realtorio':
        pass
    
    elif page == 'cadastro':
        paginas.pagina_cadastro()
#---------------------------------------------------------------------estrutura da janela---------------------------------------------------------------------#

janela = tk.Tk()

janela.title("Gestão de Pedidos")
janela.geometry("1050x650")
janela.configure(bg='#fff000')
janela.rowconfigure(0, weight=1)  # Permite que a linha 0 se expanda
janela.columnconfigure(1, weight=1) # Permite que a coluna 1 se expanda


janela.grid_rowconfigure(0, weight=1)
janela.grid_columnconfigure(1, weight=1)

style.style()

menu_frame = ttk.Frame(janela, width=200, padding=10, style="MenuFrame.TFrame")
menu_frame.grid(row=0, column=0, sticky='ns')
menu_frame.columnconfigure(0,weight=1)

try:
    logo_image = Image.open("cimed.png")
    logo_image = logo_image.resize((150, 60), Image.LANCZOS)
    logo_tk = ImageTk.PhotoImage(logo_image)
except FileNotFoundError:
    print("Erro: Logo não encontrado.")
    logo_tk = None

if logo_tk:
    logo_label = ttk.Label(menu_frame, image=logo_tk, style="TLabel")
    logo_label.pack(pady=(0, 10))

content_frame = ttk.Frame(janela,padding=5)
content_frame.grid(row=0, column=1, sticky='nsew')

for i in range(5):
    content_frame.columnconfigure(i, weight=1)

janela.grid_rowconfigure(0, weight=1)
janela.grid_columnconfigure(1, weight=1)

botoes_menu = [
    ("Consulta por Pedido", "pedidos"),
    ("Pedidos Atrasados", "atrasados"),
    ("Consulta Por Fornecedor", "fornecedor"),
    ("Relatorio", "relatorio"),
    ("Cadastro de Fornecedor", "cadastro"),
    ('Envio de Follow-Up','follow_up')
]

botao_ativo = None

def atualizar_botoes_menu(botao_clicado):
    global botao_ativo
    if botao_ativo:
        botao_ativo.configure(style="Menu.TButton")  # Volta ao estilo padrão
    botao_clicado.configure(style="ActiveMenu.TButton")  # Define estilo ativo
    botao_ativo = botao_clicado  # Atualiza o botão ativo

def criar_botao(texto, pagina):
    botao = ttk.Button(menu_frame, text=texto, style="Menu.TButton", 
                       command=lambda: [show_page(pagina), atualizar_botoes_menu(botao)])
    botao.pack(pady=(0, 5))
    return botao

for texto, pagina in botoes_menu:
    criar_botao(texto, pagina)

show_page("pedidos")

# ttk.Label(menu_frame,text='teste').pack(side = 'bottom')

janela.mainloop()